import csv
import openpyxl
from django.http import HttpResponse

def generateCsvResponse(data, filename, columns=None, column_labels=None):
    """
    Generates a CSV HttpResponse from a list of dictionaries.
    If 'columns' is provided, it dictates the order and subset of keys.
    If 'column_labels' is provided, uses human-readable labels as CSV headers.
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

    if not data:
        return response

    if not columns:
        columns = list(data[0].keys())

    # Build display headers: use labels if available, otherwise raw column names
    if column_labels:
        headers = [column_labels.get(col, col) for col in columns]
    else:
        headers = columns

    # Write UTF-8 BOM so Excel opens it correctly with accents
    response.write('\ufeff')

    writer = csv.writer(response, delimiter=';')
    writer.writerow(headers)
    for row in data:
        writer.writerow([
            str(row.get(col, '')) if row.get(col) is not None else ''
            for col in columns
        ])

    return response

def generateXlsxResponse(data, filename, columns=None, column_labels=None):
    """
    Generates an XLSX HttpResponse from a list of dictionaries.
    If 'column_labels' is provided, uses human-readable labels as sheet headers.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Export"

    if not data:
        workbook.save(response)
        return response

    if not columns:
        columns = list(data[0].keys())

    # Build display headers: use labels if available, otherwise raw column names
    if column_labels:
        headers = [column_labels.get(col, col) for col in columns]
    else:
        headers = columns

    # Write headers
    worksheet.append(headers)

    import datetime

    def _clean_for_excel(val):
        if val is None:
            return ''
        if isinstance(val, (int, float)):
            return val
        if isinstance(val, datetime.datetime):
            if val.tzinfo is not None:
                return val.replace(tzinfo=None)
            return val
        if isinstance(val, (datetime.date, datetime.time)):
            return val
        return str(val)

    # Write data rows
    for row in data:
        filtered_row = [_clean_for_excel(row.get(col)) for col in columns]
        worksheet.append(filtered_row)

    # Adjust column widths so dates/long texts don't show up as '######'
    from openpyxl.utils import get_column_letter
    for col_idx, column_cells in enumerate(worksheet.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column_cells:
            if cell.value:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
        adjusted_width = min(max_length + 2, 50) # Cap maximum width at 50 to avoid giant columns
        # Ensure a minimum width so headers don't get squished either
        adjusted_width = max(adjusted_width, 10)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(response)
    return response
